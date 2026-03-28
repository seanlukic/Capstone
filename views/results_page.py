from copy import copy
from io import BytesIO
from pathlib import Path
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from template_parser import _clean_text, table_diversity_score


OUTPUT_TEMPLATE_CANDIDATES = [
    Path(__file__).resolve().parent.parent / "assets" / "Model_Output_Template_3.25.xlsx",
    Path(__file__).resolve().parent.parent / "assets" / "Model_Output_Template_3.24.xlsx",
]
OUTPUT_DOWNLOAD_NAME = "Model_Output_Template_3.25.xlsx"


def _get_output_template_path() -> Path | None:
    for path in OUTPUT_TEMPLATE_CANDIDATES:
        if path.exists():
            return path
    return None


def _normalized_table_diversity_score(table_df, diversity_cols: list[str]) -> float:
    characteristic_count = max(1, len(diversity_cols))
    participant_count = max(1, len(table_df))
    raw_score = float(table_diversity_score(table_df, diversity_cols))
    return raw_score / characteristic_count / participant_count


def _calculate_total_balance_std_dev(schedule_results, participant_results, diversity_cols: list[str]) -> float:
    all_rounds = sorted(schedule_results["Round"].unique().tolist())
    round_average_scores = []

    for round_number in all_rounds:
        round_df = schedule_results[schedule_results["Round"] == round_number]
        table_numbers = sorted(round_df["Table"].unique().tolist())
        table_scores = []
        for table_number in table_numbers:
            table_rows = round_df[round_df["Table"] == table_number]
            person_indices = table_rows["Person_Index"].astype(int).tolist()
            table_df = participant_results.iloc[person_indices]
            table_scores.append(_normalized_table_diversity_score(table_df, diversity_cols))

        if table_scores:
            round_average_scores.append(sum(table_scores) / len(table_scores))

    if not round_average_scores:
        return 0.0

    if len(round_average_scores) == 1:
        std_dev = 0.0
    else:
        mean_score = sum(round_average_scores) / len(round_average_scores)
        variance = sum((score - mean_score) ** 2 for score in round_average_scores) / len(round_average_scores)
        std_dev = variance ** 0.5

    return max(0.0, float(std_dev))


def _total_balance_status(total_balance_std_dev: float) -> str:
    if total_balance_std_dev <= 0.05:
        return "Optimal"
    if total_balance_std_dev <= 0.10:
        return "Good"
    if total_balance_std_dev <= 0.20:
        return "Okay"
    return "Bad"


def _find_row_by_first_cell(worksheet, label: str) -> int | None:
    target = _clean_text(label).lower()
    for row_idx in range(1, worksheet.max_row + 1):
        if _clean_text(worksheet.cell(row=row_idx, column=1).value).lower() == target:
            return row_idx
    return None


def _copy_cell_format(source_cell, target_cell) -> None:
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def _clear_sheet_rows(worksheet, start_row: int) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_row >= start_row:
            worksheet.unmerge_cells(str(merged_range))

    clear_end_row = max(worksheet.max_row, start_row + 500)
    clear_end_col = max(worksheet.max_column, 200)
    for row_idx in range(start_row, clear_end_row + 1):
        for col_idx in range(1, clear_end_col + 1):
            worksheet.cell(row=row_idx, column=col_idx).value = None


def _ordered_trait_keys(
    participant_results,
    characteristics: list[str],
    trait_targets: dict,
    trait_max_allowed: dict,
    trait_min_required: dict,
) -> list[tuple[str, str]]:
    ordered_keys = []
    seen = set()

    for source in (trait_targets, trait_max_allowed, trait_min_required):
        for key in source.keys():
            if not isinstance(key, tuple) or len(key) != 2:
                continue
            normalized = (_clean_text(key[0]), _clean_text(key[1]))
            if normalized[0] and normalized[1] and normalized not in seen:
                seen.add(normalized)
                ordered_keys.append(normalized)

    for characteristic in characteristics:
        characteristic_name = _clean_text(characteristic)
        if not characteristic_name or characteristic_name not in participant_results.columns:
            continue
        values = (
            participant_results[characteristic_name]
            .dropna()
            .astype(str)
            .str.strip()
        )
        values = values[values.ne("")]
        for trait in values.drop_duplicates().tolist():
            key = (characteristic_name, trait)
            if key not in seen:
                seen.add(key)
                ordered_keys.append(key)

    return ordered_keys


def _trait_goal_text(
    key: tuple[str, str],
    trait_targets: dict,
    trait_max_allowed: dict,
    trait_min_required: dict,
) -> str:
    min_required = trait_min_required.get(key)
    max_allowed = trait_max_allowed.get(key)
    target = trait_targets.get(key)

    if min_required is not None and max_allowed is not None:
        if float(min_required) == float(max_allowed):
            return f"Goal: {float(min_required):g}"
        return f"Goal: {float(min_required):g}-{float(max_allowed):g}"
    if min_required is not None:
        return f"Min: {float(min_required):g}"
    if max_allowed is not None:
        return f"Max: {float(max_allowed):g}"
    if target is not None:
        return f"Target: {float(target):g}"
    return "No goal configured"


def _trait_deviation(
    actual_count: int,
    key: tuple[str, str],
    trait_targets: dict,
    trait_max_allowed: dict,
    trait_min_required: dict,
) -> float:
    min_required = trait_min_required.get(key)
    max_allowed = trait_max_allowed.get(key)
    target = trait_targets.get(key)

    lower = None if min_required is None else float(min_required)
    upper = None if max_allowed is None else float(max_allowed)
    if lower is None and upper is None and target is not None:
        lower = float(target)
        upper = float(target)

    deviation = 0.0
    if lower is not None and actual_count < lower:
        deviation += lower - float(actual_count)
    if upper is not None and actual_count > upper:
        deviation += float(actual_count) - upper
    return deviation


def _write_trait_deviation_view(
    workbook,
    participant_results,
    schedule_results,
    characteristics: list[str],
    trait_targets: dict,
    trait_max_allowed: dict,
    trait_min_required: dict,
) -> None:
    if "Trait Deviation View" not in workbook.sheetnames:
        raise ValueError("Output template is missing the 'Trait Deviation View' sheet.")

    worksheet = workbook["Trait Deviation View"]
    start_row = 16
    _clear_sheet_rows(worksheet, start_row)

    dark_header_template = worksheet["A5"]
    light_header_template = worksheet["A9"]
    body_label_template = worksheet["A18"]
    body_value_template = worksheet["B18"]
    total_label_template = worksheet["A24"]
    total_value_template = worksheet["C24"]

    thin_side = Side(style="thin", color="FF000000")
    outlined_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ordered_traits = _ordered_trait_keys(
        participant_results,
        characteristics,
        trait_targets,
        trait_max_allowed,
        trait_min_required,
    )

    if not ordered_traits:
        empty_title = worksheet.cell(row=start_row, column=1)
        _copy_cell_format(dark_header_template, empty_title)
        empty_title.value = "No trait data available"
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

        empty_message = worksheet.cell(row=start_row + 1, column=1)
        _copy_cell_format(body_label_template, empty_message)
        empty_message.value = "No characteristic-trait combinations were found in the solver output."
        worksheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=4)
        return

    rounds = sorted(schedule_results["Round"].dropna().astype(int).unique().tolist())
    max_col = 1 + (2 * len(ordered_traits))

    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    title_cell = worksheet.cell(row=start_row, column=1)
    _copy_cell_format(dark_header_template, title_cell)
    title_cell.value = "Trait counts and deviations by round and table"

    worksheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=max_col)
    description_cell = worksheet.cell(row=start_row + 1, column=1)
    _copy_cell_format(body_label_template, description_cell)
    description_cell.value = (
        "Count shows the number of assigned participants with each trait. "
        "Deviation is 0 inside the configured min/max range and uses the target as an exact goal when no range is configured."
    )

    worksheet.column_dimensions["A"].width = 14
    for trait_idx in range(len(ordered_traits)):
        count_col = 2 + (trait_idx * 2)
        deviation_col = count_col + 1
        worksheet.column_dimensions[get_column_letter(count_col)].width = 16
        worksheet.column_dimensions[get_column_letter(deviation_col)].width = 12

    current_row = start_row + 3
    overall_deviations = {key: 0.0 for key in ordered_traits}

    for round_number in rounds:
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
        round_title_cell = worksheet.cell(row=current_row, column=1)
        _copy_cell_format(dark_header_template, round_title_cell)
        round_title_cell.value = f"Round {round_number}"
        current_row += 1

        table_header_cell = worksheet.cell(row=current_row, column=1)
        _copy_cell_format(light_header_template, table_header_cell)
        table_header_cell.value = "Table"
        table_header_cell.border = outlined_border
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)

        for trait_idx, key in enumerate(ordered_traits):
            count_col = 2 + (trait_idx * 2)
            deviation_col = count_col + 1

            worksheet.merge_cells(
                start_row=current_row,
                start_column=count_col,
                end_row=current_row,
                end_column=deviation_col,
            )
            trait_header_cell = worksheet.cell(row=current_row, column=count_col)
            _copy_cell_format(light_header_template, trait_header_cell)
            trait_header_cell.value = f"{key[0]}: {key[1]}"
            trait_header_cell.border = outlined_border

            count_header_cell = worksheet.cell(row=current_row + 1, column=count_col)
            _copy_cell_format(light_header_template, count_header_cell)
            count_header_cell.value = "Count"
            count_header_cell.border = outlined_border

            deviation_header_cell = worksheet.cell(row=current_row + 1, column=deviation_col)
            _copy_cell_format(light_header_template, deviation_header_cell)
            deviation_header_cell.value = "Deviation"
            deviation_header_cell.border = outlined_border

        current_row += 2
        round_df = schedule_results[schedule_results["Round"] == round_number]
        tables = sorted(round_df["Table"].dropna().astype(int).unique().tolist())
        round_deviations = {key: 0.0 for key in ordered_traits}

        for table_number in tables:
            table_label_cell = worksheet.cell(row=current_row, column=1)
            _copy_cell_format(body_label_template, table_label_cell)
            table_label_cell.value = f"Table {table_number}"
            table_label_cell.border = outlined_border

            table_rows = schedule_results[
                (schedule_results["Round"] == round_number) & (schedule_results["Table"] == table_number)
            ]
            person_indices = table_rows["Person_Index"].astype(int).tolist()
            table_df = participant_results.iloc[person_indices] if person_indices else participant_results.iloc[0:0]

            for trait_idx, key in enumerate(ordered_traits):
                count_col = 2 + (trait_idx * 2)
                deviation_col = count_col + 1
                characteristic, trait = key
                if characteristic in table_df.columns:
                    actual_count = int(table_df[characteristic].astype(str).str.strip().eq(trait).sum())
                else:
                    actual_count = 0
                deviation = _trait_deviation(
                    actual_count,
                    key,
                    trait_targets,
                    trait_max_allowed,
                    trait_min_required,
                )
                count_cell = worksheet.cell(row=current_row, column=count_col)
                _copy_cell_format(body_value_template, count_cell)
                count_cell.value = actual_count
                count_cell.border = outlined_border

                deviation_cell = worksheet.cell(row=current_row, column=deviation_col)
                _copy_cell_format(body_value_template, deviation_cell)
                deviation_cell.value = round(float(deviation), 4)
                deviation_cell.border = outlined_border

                round_deviations[key] += float(deviation)
                overall_deviations[key] += float(deviation)

            current_row += 1

        total_label_cell = worksheet.cell(row=current_row, column=1)
        _copy_cell_format(total_label_template, total_label_cell)
        total_label_cell.value = f"Total Dev. in R{round_number}"
        total_label_cell.border = outlined_border

        for trait_idx, key in enumerate(ordered_traits):
            count_col = 2 + (trait_idx * 2)
            deviation_col = count_col + 1

            total_count_cell = worksheet.cell(row=current_row, column=count_col)
            _copy_cell_format(total_value_template, total_count_cell)
            total_count_cell.value = None
            total_count_cell.border = outlined_border

            total_deviation_cell = worksheet.cell(row=current_row, column=deviation_col)
            _copy_cell_format(total_value_template, total_deviation_cell)
            total_deviation_cell.value = round(round_deviations[key], 4)
            total_deviation_cell.border = outlined_border

        current_row += 2

    total_deviation_all_traits = sum(overall_deviations.values())
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    summary_title_cell = worksheet.cell(row=current_row, column=1)
    _copy_cell_format(dark_header_template, summary_title_cell)
    summary_title_cell.value = "Overall trait deviation summary"
    current_row += 1

    summary_headers = ["Trait", "Total Deviation", "Share of Total", "Goal"]
    for col_idx, header_text in enumerate(summary_headers, start=1):
        header_cell = worksheet.cell(row=current_row, column=col_idx)
        _copy_cell_format(light_header_template, header_cell)
        header_cell.value = header_text
        header_cell.border = outlined_border
    current_row += 1

    for key in ordered_traits:
        overall_dev = float(overall_deviations[key])
        share = 0.0 if total_deviation_all_traits == 0 else overall_dev / total_deviation_all_traits
        label_cell = worksheet.cell(row=current_row, column=1)
        _copy_cell_format(body_label_template, label_cell)
        label_cell.value = f"{key[0]}: {key[1]}"
        label_cell.border = outlined_border

        deviation_total_cell = worksheet.cell(row=current_row, column=2)
        _copy_cell_format(body_value_template, deviation_total_cell)
        deviation_total_cell.value = round(overall_dev, 4)
        deviation_total_cell.border = outlined_border

        share_cell = worksheet.cell(row=current_row, column=3)
        _copy_cell_format(body_value_template, share_cell)
        share_cell.value = share
        share_cell.number_format = "0.0%"
        share_cell.border = outlined_border

        goal_cell = worksheet.cell(row=current_row, column=4)
        _copy_cell_format(body_value_template, goal_cell)
        goal_cell.value = _trait_goal_text(
            key,
            trait_targets,
            trait_max_allowed,
            trait_min_required,
        )
        goal_cell.border = outlined_border

        current_row += 1


def _build_output_workbook(
    display_schedule,
    total_balance_std_dev: float,
    participant_results,
    schedule_results,
    characteristics: list[str],
    trait_targets: dict,
    trait_max_allowed: dict,
    trait_min_required: dict,
):
    template_path = _get_output_template_path()
    if template_path is None:
        raise FileNotFoundError(
            "Could not find the packaged output template in the repo."
        )

    workbook = load_workbook(template_path)
    if "Current Assignments" not in workbook.sheetnames:
        raise ValueError("Output template is missing the 'Current Assignments' sheet.")

    worksheet = workbook["Current Assignments"]
    expected_columns = display_schedule.columns.tolist()
    header_row = None
    header_columns = {}

    for row_idx in range(1, worksheet.max_row + 1):
        row_values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
        row_headers = {_clean_text(value): col_idx for col_idx, value in enumerate(row_values, start=1) if _clean_text(value)}
        if expected_columns and all(column in row_headers for column in expected_columns):
            header_row = row_idx
            header_columns = row_headers
            break

    if header_row is None:
        raise ValueError(
            "Could not find the assignment header row in 'Current Assignments'."
        )

    max_clear_row = max(worksheet.max_row, header_row + len(display_schedule) + 50)
    for row_idx in range(header_row + 1, max_clear_row + 1):
        for column in expected_columns:
            worksheet.cell(row=row_idx, column=header_columns[column]).value = None

    for row_offset, (_, row) in enumerate(display_schedule.iterrows(), start=1):
        for column in expected_columns:
            worksheet.cell(row=header_row + row_offset, column=header_columns[column]).value = row[column]

    if "Total Balance Score" not in workbook.sheetnames:
        raise ValueError("Output template is missing the 'Total Balance Score' sheet.")

    score_sheet = workbook["Total Balance Score"]
    value_row = _find_row_by_first_cell(score_sheet, "Value")
    status_row = _find_row_by_first_cell(score_sheet, "Status")
    if value_row is None or status_row is None:
        raise ValueError("Could not find the 'Value' and 'Status' rows in 'Total Balance Score'.")

    score_sheet.cell(row=value_row, column=2).value = round(float(total_balance_std_dev), 4)
    score_sheet.cell(row=status_row, column=2).value = _total_balance_status(float(total_balance_std_dev))

    _write_trait_deviation_view(
        workbook,
        participant_results,
        schedule_results,
        characteristics,
        trait_targets,
        trait_max_allowed,
        trait_min_required,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue(), OUTPUT_DOWNLOAD_NAME


# Step 4: Results page showing the generated group assignments, diversity scores, and allowing users to download the results as CSV.
def render(go_to) -> None:
    st.title("Run and Results")
    st.markdown(
        """
        <div class="hero">
            <h2 style="margin:0;">Generated group assignments</h2>
            <p class="app-subtitle">Review and download your grouping results.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <style>
            div[data-testid="stDownloadButton"] > button {
                background: linear-gradient(120deg, var(--accent) 0%, var(--accent-2) 100%);
                border: none !important;
                color: #ffffff !important;
                border-radius: 10px;
            }
            div[data-testid="stDownloadButton"] > button:hover {
                background: linear-gradient(120deg, var(--accent) 0%, var(--accent-2) 100%);
                color: #ffffff !important;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

    participant_results = st.session_state.get("participant_results")
    schedule_results = st.session_state.get("schedule_results")
    diversity_cols = st.session_state.get("characteristics", [])
    event_setup = st.session_state.get("event_setup", {})
    trait_targets = st.session_state.get("trait_targets", {})
    trait_max_allowed = st.session_state.get("trait_max_allowed", {})
    trait_min_required = st.session_state.get("trait_min_required", {})

    if participant_results is None or schedule_results is None:
        st.error("No grouping results found. Go back and click Generate Groupings.")
        st.stop()

    total_balance_std_dev = _calculate_total_balance_std_dev(schedule_results, participant_results, diversity_cols)

    round_count = int(event_setup.get("number_of_rounds", 3))
    participant_label_col = "Name" if "Name" in participant_results.columns else "Participant_ID"
    round_table_cols = [f"Round_{r}_Table" for r in range(1, round_count + 1)]
    schedule_cols = [participant_label_col, *round_table_cols]
    available_schedule_cols = [col for col in schedule_cols if col in participant_results.columns]

    sort_cols = [col for col in round_table_cols if col in participant_results.columns]
    if participant_label_col in participant_results.columns:
        sort_cols.append(participant_label_col)

    if sort_cols:
        display_schedule = participant_results.sort_values(sort_cols)[available_schedule_cols].reset_index(drop=True)
    else:
        display_schedule = participant_results[available_schedule_cols].reset_index(drop=True)

    if participant_label_col == "Name":
        display_schedule = display_schedule.rename(columns={"Name": "Participant_Name"})

    try:
        workbook_bytes, workbook_name = _build_output_workbook(
            display_schedule,
            total_balance_std_dev,
            participant_results,
            schedule_results,
            diversity_cols,
            trait_targets,
            trait_max_allowed,
            trait_min_required,
        )
    except Exception as exc:
        st.error(f"Could not build Excel output: {exc}")
        workbook_bytes = None
        workbook_name = None

    if workbook_bytes is not None:
        st.download_button(
            "Download Group Assignments (Excel)",
            data=workbook_bytes,
            file_name=workbook_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    all_rounds = sorted(schedule_results["Round"].unique().tolist())
    for round_number in all_rounds:
        st.subheader(f"Round {round_number}")
        round_df = schedule_results[schedule_results["Round"] == round_number]
        table_numbers = sorted(round_df["Table"].unique().tolist())
        cols = st.columns(max(1, min(3, len(table_numbers))))

        for idx, table_number in enumerate(table_numbers):
            table_rows = round_df[round_df["Table"] == table_number]
            person_indices = table_rows["Person_Index"].astype(int).tolist()
            table_df = participant_results.iloc[person_indices]
            score = table_diversity_score(table_df, diversity_cols)

            with cols[idx % len(cols)]:
                with st.container(border=True):
                    st.markdown(f"**Table {table_number}**")
                    st.caption(f"Diversity score: {score}")
                    for _, person_row in table_df.iterrows():
                        person_name = _clean_text(person_row.get("Name", ""))
                        if person_name:
                            st.write(f"- {person_name}")

    left, right = st.columns(2)
    with left:
        if st.button("Back to Participants"):
            go_to(3)
    with right:
        st.info("If needed, go back and change participant data to regenerate groups.")
